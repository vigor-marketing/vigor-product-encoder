# -*- coding: utf-8 -*-
"""
Vigor 编码器 - 本地服务
- 数据保存在本机磁盘 data.json 文件（可备份、可迁移）
- 浏览器访问 http://127.0.0.1:8765
- 无需安装任何依赖（Python 标准库）
"""
import base64
import http.server
import io
import json
import os
import re
import socket
import socketserver
import sys
import threading
import time

try:
    import openpyxl
except ImportError:
    openpyxl = None

BASE = os.path.dirname(os.path.abspath(__file__))
PUBLIC = os.path.join(BASE, 'public')
UPLOAD_DIR = os.path.join(BASE, 'uploads')
TEMPLATE_DIR = os.path.join(BASE, 'template')
TEMPLATE_FILE = os.path.join(TEMPLATE_DIR, '产品导入模版.xls')
DATA_FILE = os.path.join(BASE, 'data.json')
PORT = 8765
WRITE_LOCK = threading.Lock()
ALLOWED_EXT = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp',
               '.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx',
               '.txt', '.csv', '.zip', '.rar', '.7z'}
MIME_MAP = {
    '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg', '.png': 'image/png',
    '.gif': 'image/gif', '.webp': 'image/webp', '.bmp': 'image/bmp',
    '.pdf': 'application/pdf', '.txt': 'text/plain', '.csv': 'text/csv',
    '.doc': 'application/msword', '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    '.xls': 'application/vnd.ms-excel', '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
}


def get_lan_ip():
    """获取本机局域网 IP，供其他电脑访问"""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.settimeout(1)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return None


def load_store():
    """读取磁盘数据文件；不存在或损坏时返回空结构"""
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                store = json.load(f)
            if isinstance(store, dict):
                if not isinstance(store.get('data'), list):
                    store['data'] = []
                if not isinstance(store.get('saved'), list):
                    store['saved'] = []
                return store
        except Exception:
            pass
    return {"data": [], "saved": []}


def save_store(store):
    """写入磁盘数据文件（先写临时文件再替换，避免中途断电损坏）"""
    with WRITE_LOCK:
        tmp = DATA_FILE + '.tmp'
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump(store, f, ensure_ascii=False, indent=1)
        os.replace(tmp, DATA_FILE)


def build_import_file(rows):
    """基于 CRM 导入模板填充数据，生成可直接导入的 .xls 文件字节"""
    if openpyxl is None:
        raise RuntimeError('缺少 openpyxl，请运行: pip install openpyxl')
    if not os.path.exists(TEMPLATE_FILE):
        raise RuntimeError('缺少导入模板，请将 产品导入模版.xls 放入 template 文件夹')
    # 模板是 xlsx 容器伪装 .xls，用 BytesIO 绕过扩展名检查
    with open(TEMPLATE_FILE, 'rb') as f:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()))
    ws = wb.active
    # 清除旧数据行（第 2 行起，最多清 2000 行）
    for r in range(2, 2002):
        for c in range(1, 8):
            ws.cell(row=r, column=c).value = None
    # 写入数据
    for idx, item in enumerate(rows):
        r = idx + 2
        ws.cell(row=r, column=1, value=item.get('code', ''))
        ws.cell(row=r, column=2, value=item.get('cat', ''))
        ws.cell(row=r, column=3, value=item.get('unit', '件'))
        ws.cell(row=r, column=4, value=item.get('enabled', '是'))
        ws.cell(row=r, column=5, value=item.get('cn', ''))
        ws.cell(row=r, column=6, value=item.get('en', ''))
        ws.cell(row=r, column=7, value=item.get('note', ''))
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def do_import(rows):
    """批量导入产品信息（标准→子类→产品→参数组→选项 层级）。
    rows: [{api_cn,api_en,cat_cn,cat_en,prod_cn,prod_en,params:[{name_cn,name_en,opts:[[code,desc],...]}]}]
    编码全部自动生成（标准/分类 A~Z，产品 AA~ZZ），按中文名称匹配已有实体。
    返回统计 dict。
    """
    store = load_store()
    data = store.setdefault('data', [])
    stats = {'api': 0, 'cat': 0, 'prod': 0, 'pg': 0, 'opt': 0, 'errors': []}

    def next_code(cs, length):
        mx = -1
        for c in cs:
            if isinstance(c, str) and len(c) == length and c.isalpha() and c.isupper():
                n = 0
                for ch in c:
                    n = n * 26 + (ord(ch) - 65)
                if n > mx:
                    mx = n
        nxt = mx + 1
        if length == 1:
            if nxt >= 26:
                return None
            return chr(65 + nxt)
        else:
            if nxt >= 676:
                return None
            return chr(65 + nxt // 26) + chr(65 + nxt % 26)

    def match(item, cn, en):
        """按中英文名匹配（中文优先，其次英文）"""
        if not cn and not en:
            return False
        if cn and (item.get('cname') == cn or item.get('name') == cn):
            return True
        if en and (item.get('name') == en or item.get('cname') == en):
            return True
        return False

    for row in rows:
        try:
            api_cn = str(row.get('api_cn', '') or '').strip()
            api_en = str(row.get('api_en', '') or '').strip()
            cat_cn = str(row.get('cat_cn', '') or '').strip()
            cat_en = str(row.get('cat_en', '') or '').strip()
            prod_cn = str(row.get('prod_cn', '') or '').strip()
            prod_en = str(row.get('prod_en', '') or '').strip()
            params = row.get('params') or []
            if not (api_cn or api_en):
                raise ValueError('缺少产品标准名称')

            # 标准（按名称匹配，找不到则自动生成编码 A~Z）
            api = None
            for a in data:
                if match(a, api_cn, api_en):
                    api = a
                    break
            if api is None:
                api_code = next_code([a.get('code', '') for a in data], 1)
                if not api_code:
                    raise ValueError('标准编码已满(A~Z)')
                api = {'code': api_code, 'name': api_en or api_cn, 'cname': api_cn or api_en, 'categories': []}
                data.append(api)
                stats['api'] += 1

            # 分类（子类）
            cats = api['categories']
            cat = None
            for c in cats:
                if match(c, cat_cn, cat_en):
                    cat = c
                    break
            if cat is None:
                cat_code = next_code([c.get('code', '') for c in cats], 1)
                if not cat_code:
                    raise ValueError('分类编码已满(A~Z)')
                cat = {'code': cat_code, 'name': cat_en or cat_cn, 'cname': cat_cn or cat_en, 'products': []}
                cats.append(cat)
                stats['cat'] += 1

            # 产品
            prods = cat['products']
            prod = None
            for p in prods:
                if match(p, prod_cn, prod_en):
                    prod = p
                    break
            if prod is None:
                prod_code = next_code([p.get('code', '') for p in prods], 2)
                if not prod_code:
                    raise ValueError('产品编码已满(AA~ZZ)')
                prod = {'code': prod_code, 'name': prod_en or prod_cn, 'cname': prod_cn or prod_en, 'paramGroups': []}
                prods.append(prod)
                stats['prod'] += 1

            # 参数组（K1~K7）+ 选项
            pgs = prod.setdefault('paramGroups', [])
            for p in params:
                label = str(p.get('label', '') or '').strip()
                # 新版导入列将“参数分类名称”与“参数选项名称”分开保存；旧版仍回退 name_cn/name_en。
                pcn = str(p.get('group_name_cn', p.get('name_cn', '')) or '').strip()
                pen = str(p.get('group_name_en', p.get('name_en', '')) or '').strip()
                opts = p.get('opts') or []
                code = str(p.get('code', '') or '').strip()
                desc = str(p.get('desc', '') or '').strip()
                # 新格式：编码 + 中英描述（按 K 编号匹配参数组，同组多行合并多选项）
                if label and (code or desc):
                    pg = None
                    for g in pgs:
                        if g.get('label') == label:
                            pg = g
                            break
                    if pg is None:
                        pg = {'label': label, 'name': '', 'name_en': '', 'options': []}
                        pgs.append(pg)
                        stats['pg'] += 1
                    # 允许新版导入修复历史错误：只有空值或“中英文被写成同一个值”时才替换，保留人工维护的有效翻译。
                    old_name = str(pg.get('name', '') or '').strip()
                    old_name_en = str(pg.get('name_en', '') or '').strip()
                    if pcn and (not old_name or old_name == old_name_en):
                        pg['name'] = pcn
                    if pen and (not old_name_en or old_name_en == old_name or old_name_en == pcn):
                        pg['name_en'] = pen
                    existing = None
                    for x in pg['options']:
                        if x.get('code') == code:
                            existing = x
                            break
                    if existing is None and code:
                        pg['options'].append({'code': code, 'desc': desc, 'desc_en': p.get('desc_en', '') or ''})
                        stats['opt'] += 1
                    elif existing is not None:
                        # 同编码已存在：补全缺失的中英文描述
                        if desc and not existing.get('desc'):
                            existing['desc'] = desc
                        new_en = str(p.get('desc_en', '') or '').strip()
                        if new_en and not existing.get('desc_en'):
                            existing['desc_en'] = new_en
                    continue
                # 旧格式兼容：名称列当组名，选项文本拆分
                if not (pcn or pen):
                    continue
                pg = None
                for g in pgs:
                    if g.get('name') == pcn or (pen and g.get('name') == pen):
                        pg = g
                        break
                if pg is None:
                    pg_label = 'K%d' % (len(pgs) + 1)
                    pg = {'label': pg_label, 'name': pcn or pen, 'options': []}
                    pgs.append(pg)
                    stats['pg'] += 1
                for o in opts:
                    oc, od = o
                    if not oc:
                        continue
                    exists = any(x.get('code') == oc or x.get(0) == oc for x in pg['options'])
                    if not exists:
                        pg['options'].append({'code': oc, 'desc': od})
                        stats['opt'] += 1
        except Exception as e:
            stats['errors'].append('行: %s' % e)

    save_store(store)  # save_store 内部自带写锁，这里不能再加锁（Lock 不可重入）
    return stats


class Handler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=PUBLIC, **kwargs)

    def _send_json(self, obj, status=200):
        body = json.dumps(obj, ensure_ascii=False).encode('utf-8')
        self.send_response(status)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Content-Length', str(len(body)))
        self.send_header('Cache-Control', 'no-store')
        self.end_headers()
        self.wfile.write(body)

    def end_headers(self):
        # 禁用缓存，确保页面每次都是最新（多人修改后无需强刷）
        self.send_header('Cache-Control', 'no-store, no-cache, must-revalidate')
        super().end_headers()

    def do_GET(self):
        if self.path == '/api/data':
            self._send_json(load_store())
            return
        # 上传文件静态服务：/uploads/文件名
        if self.path.startswith('/uploads/'):
            rel = self.path[len('/uploads/'):]
            if '..' in rel or '/' in rel:
                self._send_json({"ok": False, "err": "bad path"}, 404)
                return
            fpath = os.path.join(UPLOAD_DIR, rel)
            if os.path.isfile(fpath):
                ext = os.path.splitext(fpath)[1].lower()
                ctype = MIME_MAP.get(ext, 'application/octet-stream')
                with open(fpath, 'rb') as f:
                    body = f.read()
                self.send_response(200)
                self.send_header('Content-Type', ctype)
                self.send_header('Content-Length', str(len(body)))
                self.end_headers()
                self.wfile.write(body)
            else:
                self._send_json({"ok": False, "err": "file not found"}, 404)
            return
        super().do_GET()

    def do_POST(self):
        if self.path == '/api/data':
            try:
                length = int(self.headers.get('Content-Length', 0))
                raw = self.rfile.read(length)
                store = json.loads(raw.decode('utf-8'))
                if not isinstance(store, dict):
                    raise ValueError('data must be an object')
                if not isinstance(store.get('data'), list):
                    store['data'] = []
                if not isinstance(store.get('saved'), list):
                    store['saved'] = []
                save_store(store)
                self._send_json({"ok": True})
            except Exception as e:
                self._send_json({"ok": False, "err": str(e)}, 400)
            return
        if self.path == '/api/upload':
            try:
                length = int(self.headers.get('Content-Length', 0))
                raw = self.rfile.read(length)
                req = json.loads(raw.decode('utf-8'))
                name = str(req.get('name', 'file'))
                b64 = str(req.get('data', ''))
                if not b64:
                    raise ValueError('no file data')
                # 文件名安全化：只保留字母数字中文下划线连字符
                base = re.sub(r'[^\w\u4e00-\u9fff-]', '_', os.path.splitext(name)[0])[:50] or 'file'
                ext = os.path.splitext(name)[1].lower()
                if ext not in ALLOWED_EXT:
                    ext = '.bin'
                fname = '%s_%d%s' % (base, int(time.time() * 1000), ext)
                os.makedirs(UPLOAD_DIR, exist_ok=True)
                fpath = os.path.join(UPLOAD_DIR, fname)
                with open(fpath, 'wb') as f:
                    f.write(base64.b64decode(b64))
                self._send_json({"ok": True, "url": '/uploads/' + fname, "name": fname})
            except Exception as e:
                self._send_json({"ok": False, "err": str(e)}, 400)
            return
        if self.path == '/api/export':
            try:
                length = int(self.headers.get('Content-Length', 0))
                raw = self.rfile.read(length)
                req = json.loads(raw.decode('utf-8'))
                rows = req.get('rows', [])
                if not isinstance(rows, list) or not rows:
                    raise ValueError('rows must be a non-empty array')
                body = build_import_file(rows)
                print('[export] building ok, size=%d rows=%d' % (len(body), len(rows)), file=sys.stderr)
                # 注意：HTTP 头只支持 latin-1，文件名用 ASCII（前端下载时用 a.download 指定中文名）
                self.send_response(200)
                self.send_header('Content-Type', 'application/vnd.ms-excel; charset=utf-8')
                self.send_header('Content-Length', str(len(body)))
                self.send_header('Content-Disposition', 'attachment; filename="Vigor_export.xls"')
                self.end_headers()
                self.wfile.write(body)
                self.wfile.flush()
                print('[export] body written', file=sys.stderr)
            except Exception as e:
                import traceback
                traceback.print_exc(file=sys.stderr)
                self._send_json({"ok": False, "err": str(e)}, 400)
            return
        if self.path == '/api/import':
            try:
                length = int(self.headers.get('Content-Length', 0))
                raw = self.rfile.read(length)
                req = json.loads(raw.decode('utf-8'))
                rows = req.get('rows', [])
                if not isinstance(rows, list) or not rows:
                    raise ValueError('rows must be a non-empty array')
                stats = do_import(rows)
                self._send_json({"ok": True, "stats": stats})
            except Exception as e:
                import traceback
                traceback.print_exc(file=sys.stderr)
                self._send_json({"ok": False, "err": str(e)}, 400)
            return
        if self.path == '/api/import_file':
            try:
                length = int(self.headers.get('Content-Length', 0))
                raw = self.rfile.read(length)
                req = json.loads(raw.decode('utf-8'))
                name = str(req.get('name', 'import.xlsx'))
                b64 = str(req.get('data', ''))
                if not b64:
                    raise ValueError('no file data')
                tmp = os.path.join(BASE, '_tmp_import.xlsx')
                with open(tmp, 'wb') as f:
                    f.write(base64.b64decode(b64))
                try:
                    wb = openpyxl.load_workbook(tmp)
                finally:
                    try:
                        if os.path.exists(tmp):
                            os.remove(tmp)
                    except Exception:
                        pass  # 沙箱环境可能禁止删除，临时文件同名覆盖即可
                ws = wb['产品导入'] if '产品导入' in wb.sheetnames else wb.active
                rows = []
                # 新版 41 列：基础 6 列 + K1-K7 每组 5 列（分类中/分类英/编码/选项中/选项英）。
                # 旧版 27 列仍可导入，避免已发出的历史模板失效。
                # 以第 7 列是否为新版的“参数1分类名称(中)”精确识别，避免旧表附带格式列时错位。
                first_param_header = str(ws.cell(row=1, column=7).value or '').strip()
                is_new_layout = first_param_header == '参数1分类名称(中)'
                last_col = 41 if is_new_layout else 27
                for r in range(2, ws.max_row + 1):
                    vals = [ws.cell(row=r, column=c).value for c in range(1, last_col + 1)]
                    if all(v is None or str(v).strip() == '' for v in vals):
                        continue
                    # 跳过模板示例行/提示行：黄色填充(FEF3C7) 或 文本含"示例/黄色/删除"
                    cell1 = ws.cell(row=r, column=1)
                    try:
                        fl = cell1.fill
                        fg = ''
                        if fl and fl.patternType:
                            fg = str(fl.fgColor.rgb if fl.fgColor else '')
                        if 'FEF3C7' in fg:
                            continue
                    except Exception:
                        pass
                    row_txt = ''.join(str(v) for v in vals if v is not None)
                    if '示例' in row_txt or '黄色' in row_txt or '删除后填写' in row_txt:
                        continue
                    def s(v):
                        return '' if v is None else str(v).strip()
                    params = []
                    for i in range(7):
                        label = 'K%d' % (i + 1)
                        if is_new_layout:
                            # 新版顺序：参数分类名称(中)、参数分类名称(英)、参数编码、参数名称(中)、参数名称(英)
                            base = 6 + i * 5
                            group_cn = s(vals[base])
                            group_en = s(vals[base + 1])
                            code_txt = s(vals[base + 2])
                            name_cn = s(vals[base + 3])
                            name_en = s(vals[base + 4])
                            if group_cn or group_en or code_txt or name_cn or name_en:
                                params.append({'label': label, 'group_name_cn': group_cn, 'group_name_en': group_en, 'code': code_txt, 'desc': name_cn, 'desc_en': name_en})
                            continue
                        # 旧版顺序：参数编码、参数名称(中)、参数名称(英)
                        code_txt = s(vals[6 + i * 3])
                        name_cn = s(vals[7 + i * 3])
                        name_en = s(vals[8 + i * 3])
                        if not (name_cn or name_en or code_txt):
                            continue
                        if ';' not in code_txt and '；' not in code_txt:
                            params.append({'label': label, 'code': code_txt, 'desc': name_cn, 'desc_en': name_en})
                            continue
                        opts = []
                        for part in re.split(r'[;；]', code_txt):
                            part = part.strip()
                            if not part:
                                continue
                            seg = part.split(None, 1)
                            opts.append([seg[0], seg[1] if len(seg) > 1 else ''])
                        params.append({'label': label, 'name_cn': name_cn, 'name_en': name_en, 'opts': opts})
                    rows.append({
                        'api_cn': s(vals[0]), 'api_en': s(vals[1]),
                        'cat_cn': s(vals[2]), 'cat_en': s(vals[3]),
                        'prod_cn': s(vals[4]), 'prod_en': s(vals[5]),
                        'params': params,
                    })
                stats = do_import(rows)
                self._send_json({"ok": True, "file": name, "rows": len(rows), "stats": stats})
            except Exception as e:
                import traceback
                traceback.print_exc(file=sys.stderr)
                self._send_json({"ok": False, "err": str(e)}, 400)
            return
        self._send_json({"ok": False, "err": "not found"}, 404)

    def log_message(self, fmt, *args):
        pass  # 静默日志


if __name__ == '__main__':
    os.makedirs(PUBLIC, exist_ok=True)
    socketserver.TCPServer.allow_reuse_address = True
    lan = get_lan_ip()
    print('=' * 56)
    print('  Vigor 编码器 - 局域网共享版')
    print('  ' + '-' * 50)
    print('  本机访问   : http://127.0.0.1:%d' % PORT)
    if lan:
        print('  局域网访问 : http://%s:%d' % (lan, PORT))
        print('  (把上面局域网地址发给其他同事即可访问)')
    print('  数据文件   : %s' % DATA_FILE)
    print('  上传目录   : %s' % UPLOAD_DIR)
    print('  导入模板   : %s' % TEMPLATE_FILE)
    if openpyxl is None:
        print('  !!! 缺少 openpyxl，导出功能不可用，请运行: pip install openpyxl')
    print('  ' + '-' * 50)
    print('  ★ 如果其他电脑打不开：')
    print('    Windows 防火墙弹出提示时选择【允许访问】')
    print('    或手动放行：控制面板 → Windows防火墙 → 允许应用 →')
    print('    勾选 Python 的【专用网络】和【公用网络】')
    print('  ' + '-' * 50)
    print('  关闭: 直接关掉本窗口，或按 Ctrl+C')
    print('=' * 56)
    # 绑定 0.0.0.0 允许局域网其他设备访问
    with socketserver.ThreadingTCPServer(('0.0.0.0', PORT), Handler) as httpd:
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print('\n服务已停止')
