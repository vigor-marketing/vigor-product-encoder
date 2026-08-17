# -*- coding: utf-8 -*-
"""CRM 导出：基于模板生成 7 列导入文件。"""
import io

import pytest


def test_build_import_file(server):
    pytest.importorskip("openpyxl")
    rows = [{
        "code": "AABB", "cat": "套管", "unit": "件", "enabled": "是",
        "cn": "套管和油管。套管。", "en": "API Spec 5CT Casing Casing.", "note": "备注",
    }]
    body = server.build_import_file(rows)
    assert isinstance(body, bytes) and len(body) > 0

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(body))
    ws = wb.active
    assert ws.cell(row=2, column=1).value == "AABB"
    assert ws.cell(row=2, column=2).value == "套管"
    assert ws.cell(row=2, column=3).value == "件"
    assert ws.cell(row=2, column=4).value == "是"
    assert ws.cell(row=2, column=5).value == "套管和油管。套管。"
    assert ws.cell(row=2, column=6).value == "API Spec 5CT Casing Casing."
    assert ws.cell(row=2, column=7).value == "备注"
