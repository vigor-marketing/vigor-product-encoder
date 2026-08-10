# -*- coding: utf-8 -*-
"""
Vigor 编码器 - 云端版（CloudBase 云托管）
- 数据持久化到 CloudBase 云存储（COS），容器重启不丢
- 上传文件存云存储
- 监听平台注入的 PORT
"""
import base64
import http.server
import io
import json
import os
import re
import socket
import socketserver
import sys
import threading
import time

# COS SDK（可选，缺失时降级本地磁盘）
try:
    from qcloud_cos import CosConfig, CosS3Client
    COS_OK = True
except ImportError:
    COS_OK = False

try:
    import openpyxl
except ImportError:
    openpyxl = None

BASE = os.path.dirname(os.path.abspath(__file__))
PUBLIC = os.path.join(BASE, 'public')
TEMPLATE_DIR = os.path.join(BASE, 'template')
TEMPLATE_FILE = os.path.join(TEMPLATE_DIR, '产品导入模版.xls')
DATA_FILE = os.path.join(BASE, 'data.json')
UPLOAD_DIR = os.path.join(BASE, 'uploads')
PORT = int(os.environ.get('PORT') or '3000')  # 云托管注入 PORT；缺失时默认 3000（探针端口）
WRITE_LOCK = threading.Lock()
ALLOWED_EXT = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp',
               '.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx',
               '.txt', '.csv', '.zip', '.rar', '.7z'}
MIME_MAP = {
    '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg', '.png': 'image/png',
    '.gif': 'image/gif', '.webp': 'image/webp', '.bmp': 'image/bmp',
    '.pdf': 'application/pdf', '.txt': 'text/plain', '.csv': 'text/csv',
    '.doc': 'application/msword', '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    '.xls': 'application/vnd.ms-excel', '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
}

# ===== 云存储（COS）持久化 =====
COS_BUCKET = os.environ.get('TCB_BUCKET', '6d6f-monktestcloud-d8gnzlwaw449aa8b8-1459141414')
COS_REGION = os.environ.get('TCB_REGION', 'ap-shanghai')
DATA_KEY = 'vigor/data.json'
UPLOAD_PREFIX = 'vigor/uploads/'


def get_cos_client():
    """从云托管注入的环境变量获取临时密钥，创建 COS 客户端"""
    if not COS_OK:
        return None
    sid = os.environ.get('TENCENTCLOUD_SECRETID') or os.environ.get('TCB_SECRETID')
    skey = os.environ.get('TENCENTCLOUD_SECRETKEY') or os.environ.get('TCB_SECRETKEY')
    token = os.environ.get('TENCENTCLOUD_SESSIONTOKEN') or os.environ.get('TCB_SESSIONTOKEN')
    if not (sid and skey):
        print('[storage] 未找到云托管注入的临时密钥，使用本地磁盘模式', file=sys.stderr)
        return None
    try:
        config = CosConfig(Region=COS_REGION, SecretId=sid, SecretKey=skey, Token=token or None)
        return CosS3Client(config)
    except Exception as e:
        print('[storage] COS 客户端创建失败: %s' % e, file=sys.stderr)
        return None


_cos_cache = {'client': None}


def cos():
    if _cos_cache['client'] is None:
        _cos_cache['client'] = get_cos_client()
    return _cos_cache['client']


def storage_get(key):
    """从云存储读取文件内容，失败返回 None"""
    c = cos()
    if c is None:
        return None
    try:
        resp = c.get_object(Bucket=COS_BUCKET, Key=key)
        return resp['Body'].get_raw_stream().read()
    except Exception as e:
        print('[storage] 读取失败 %s: %s' % (key, e), file=sys.stderr)
        return None


def storage_put(key, data):
    """写入云存储"""
    c = cos()
    if c is None:
        return False
    try:
        c.put_object(Bucket=COS_BUCKET, Key=key, Body=data)
        return True
    except Exception as e:
        print('[storage] 写入失败 %s: %s' % (key, e), file=sys.stderr)
        return False


def storage_del(key):
    c = cos()
    if c is None:
        return
    try:
        c.delete_object(Bucket=COS_BUCKET, Key=key)
    except Exception:
        pass


# ===== 数据读写（云存储优先，本地缓存兜底） =====
def load_store():
    # 1. 云存储
    raw = storage_get(DATA_KEY)
    if raw is not None:
        try:
            store = json.loads(raw.decode('utf-8'))
            if isinstance(store, dict):
                if not isinstance(store.get('data'), list):
                    store['data'] = []
                if not isinstance(store.get('saved'), list):
                    store['saved'] = []
                # 本地缓存
                try:
                    with open(DATA_FILE, 'w', encoding='utf-8') as f:
                        json.dump(store, f, ensure_ascii=False, indent=1)
                except Exception:
                    pass
                return store
        except Exception as e:
            print('[data] 云存储数据解析失败: %s' % e, file=sys.stderr)
    # 2. 本地缓存
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                store = json.load(f)
            if isinstance(store, dict):
                if not isinstance(store.get('data'), list):
                    store['data'] = []
                if not isinstance(store.get('saved'), list):
                    store['saved'] = []
                return store
        except Exception:
            pass
    return {"data": [], "saved": []}


def save_store(store):
    """写入云存储 + 本地缓存"""
    data = json.dumps(store, ensure_ascii=False, indent=1).encode('utf-8')
    storage_put(DATA_KEY, data)
    try:
        with open(DATA_FILE, 'w', encoding='utf-8') as f:
            f.write(data.decode('utf-8'))
    except Exception:
        pass


def build_import_file(rows):
    """基于 CRM 导入模板填充数据，生成可直接导入的 .xls 文件字节"""
    if openpyxl is None:
        raise RuntimeError('缺少 openpyxl')
    if not os.path.exists(TEMPLATE_FILE):
        raise RuntimeError('缺少导入模板')
    with open(TEMPLATE_FILE, 'rb') as f:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()))
    ws = wb.active
    for r in range(2, 2002):
        for c in range(1, 8):
            ws.cell(row=r, column=c).value = None
    for idx, item in enumerate(rows):
        r = idx + 2
        ws.cell(row=r, column=1, value=item.get('code', ''))
        ws.cell(row=r, column=2, value=item.get('cat', ''))
        ws.cell(row=r, column=3, value=item.get('unit', '件'))
        ws.cell(row=r, column=4, value=item.get('enabled', '是'))
        ws.cell(row=r, column=5, value=item.get('cn', ''))
        ws.cell(row=r, column=6, value=item.get('en', ''))
        ws.cell(row=r, column=7, value=item.get('note', ''))
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def do_import(rows):
    """批量导入产品信息（标准→子类→产品→参数组→选项 层级）。编码全自动，按中文名称匹配。"""
    store = load_store()
    data = store.setdefault('data', [])
    stats = {'api': 0, 'cat': 0, 'prod': 0, 'pg': 0, 'opt': 0, 'errors': []}

    def next_code(cs, length):
        mx = -1
        for c in cs:
            if isinstance(c, str) and len(c) == length and c.isalpha() and c.isupper():
                n = 0
                for ch in c:
                    n = n * 26 + (ord(ch) - 65)
                if n > mx:
                    mx = n
        nxt = mx + 1
        if length == 1:
            if nxt >= 26:
                return None
            return chr(65 + nxt)
        else:
            if nxt >= 676:
                return None
            return chr(65 + nxt // 26) + chr(65 + nxt % 26)

    def match(item, cn, en):
        if not cn and not en:
            return False
        if cn and (item.get('cname') == cn or item.get('name') == cn):
            return True
        if en and (item.get('name') == en or item.get('cname') == en):
            return True
        return False

    for row in rows:
        try:
            api_cn = str(row.get('api_cn', '') or '').strip()
            api_en = str(row.get('api_en', '') or '').strip()
            cat_cn = str(row.get('cat_cn', '') or '').strip()
            cat_en = str(row.get('cat_en', '') or '').strip()
            prod_cn = str(row.get('prod_cn', '') or '').strip()
            prod_en = str(row.get('prod_en', '') or '').strip()
            params = row.get('params') or []
            if not (api_cn or api_en):
                raise ValueError('缺少产品标准名称')

            api = None
            for a in data:
                if match(a, api_cn, api_en):
                    api = a
                    break
            if api is None:
                api_code = next_code([a.get('code', '') for a in data], 1)
                if not api_code:
                    raise ValueError('标准编码已满(A~Z)')
                api = {'code': api_code, 'name': api_en or api_cn, 'cname': api_cn or api_en, 'categories': []}
                data.append(api)
                stats['api'] += 1

            cats = api['categories']
            cat = None
            for c in cats:
                if match(c, cat_cn, cat_en):
                    cat = c
                    break
            if cat is None:
                cat_code = next_code([c.get('code', '') for c in cats], 1)
                if not cat_code:
                    raise ValueError('分类编码已满(A~Z)')
                cat = {'code': cat_code, 'name': cat_en or cat_cn, 'cname': cat_cn or cat_en, 'products': []}
                cats.append(cat)
                stats['cat'] += 1

            prods = cat['products']
            prod = None
            for p in prods:
                if match(p, prod_cn, prod_en):
                    prod = p
                    break
            if prod is None:
                prod_code = next_code([p.get('code', '') for p in prods], 2)
                if not prod_code:
                    raise ValueError('产品编码已满(AA~ZZ)')
                prod = {'code': prod_code, 'name': prod_en or prod_cn, 'cname': prod_cn or prod_en, 'paramGroups': []}
                prods.append(prod)
                stats['prod'] += 1

            pgs = prod.setdefault('paramGroups', [])
            for p in params:
                label = str(p.get('label', '') or '').strip()
                pcn = str(p.get('name_cn', '') or '').strip()
                pen = str(p.get('name_en', '') or '').strip()
                opts = p.get('opts') or []
                code = str(p.get('code', '') or '').strip()
                desc = str(p.get('desc', '') or '').strip()
                # 新格式：编码 + 中英描述（按 K 编号匹配参数组，同组多行合并多选项）
                if label and (code or desc):
                    pg = None
                    for g in pgs:
                        if g.get('label') == label:
                            pg = g
                            break
                    if pg is None:
                        pg = {'label': label, 'name': '', 'options': []}
                        pgs.append(pg)
                        stats['pg'] += 1
                    existing = next((x for x in pg['options'] if x.get('code') == code), None)
                    if existing is None and code:
                        pg['options'].append({'code': code, 'desc': desc, 'desc_en': p.get('desc_en', '') or ''})
                        stats['opt'] += 1
                    elif existing is not None:
                        # 同编码已存在时，只补齐缺失的中英文信息
                        if desc and not existing.get('desc'):
                            existing['desc'] = desc
                        desc_en = str(p.get('desc_en', '') or '').strip()
                        if desc_en and not existing.get('desc_en'):
                            existing['desc_en'] = desc_en
                    continue
                # 旧格式兼容：名称列当组名，选项文本拆分
                if not (pcn or pen):
                    continue
                pg = None
                for g in pgs:
                    if g.get('name') == pcn or (pen and g.get('name') == pen):
                        pg = g
                        break
                if pg is None:
                    pg_label = 'K%d' % (len(pgs) + 1)
                    pg = {'label': pg_label, 'name': pcn or pen, 'options': []}
                    pgs.append(pg)
                    stats['pg'] += 1
                for o in opts:
                    oc, od = o
                    if not oc:
                        continue
                    exists = any(x.get('code') == oc or x.get(0) == oc for x in pg['options'])
                    if not exists:
                        pg['options'].append({'code': oc, 'desc': od})
                        stats['opt'] += 1
        except Exception as e:
            stats['errors'].append('行: %s' % e)

    # save_store 内部自带写锁；这里不能重复加锁（Lock 不可重入）
    save_store(store)
    return stats


class Handler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=PUBLIC, **kwargs)

    def _send_json(self, obj, status=200):
        body = json.dumps(obj, ensure_ascii=False).encode('utf-8')
        self.send_response(status)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Content-Length', str(len(body)))
        self.send_header('Cache-Control', 'no-store')
        self.end_headers()
        self.wfile.write(body)

    def end_headers(self):
        self.send_header('Cache-Control', 'no-store, no-cache, must-revalidate')
        super().end_headers()

    def do_GET(self):
        if self.path == '/api/data':
            self._send_json(load_store())
            return
        if self.path.startswith('/uploads/'):
            rel = self.path[len('/uploads/'):]
            if '..' in rel or '/' in rel:
                self._send_json({"ok": False, "err": "bad path"}, 404)
                return
            # 优先云存储，其次本地
            data = storage_get(UPLOAD_PREFIX + rel)
            if data is None:
                fpath = os.path.join(UPLOAD_DIR, rel)
                if os.path.isfile(fpath):
                    with open(fpath, 'rb') as f:
                        data = f.read()
            if data is not None:
                ext = os.path.splitext(rel)[1].lower()
                ctype = MIME_MAP.get(ext, 'application/octet-stream')
                self.send_response(200)
                self.send_header('Content-Type', ctype)
                self.send_header('Content-Length', str(len(data)))
                self.end_headers()
                self.wfile.write(data)
            else:
                self._send_json({"ok": False, "err": "file not found"}, 404)
            return
        super().do_GET()

    def do_POST(self):
        if self.path == '/api/data':
            try:
                length = int(self.headers.get('Content-Length', 0))
                raw = self.rfile.read(length)
                store = json.loads(raw.decode('utf-8'))
                if not isinstance(store, dict):
                    raise ValueError('data must be an object')
                if not isinstance(store.get('data'), list):
                    store['data'] = []
                if not isinstance(store.get('saved'), list):
                    store['saved'] = []
                # save_store 内部已加锁，避免 Lock 不可重入导致请求卡住
                save_store(store)
                self._send_json({"ok": True})
            except Exception as e:
                self._send_json({"ok": False, "err": str(e)}, 400)
            return
        if self.path == '/api/upload':
            try:
                length = int(self.headers.get('Content-Length', 0))
                raw = self.rfile.read(length)
                req = json.loads(raw.decode('utf-8'))
                name = str(req.get('name', 'file'))
                b64 = str(req.get('data', ''))
                if not b64:
                    raise ValueError('no file data')
                base = re.sub(r'[^\w\u4e00-\u9fff-]', '_', os.path.splitext(name)[0])[:50] or 'file'
                ext = os.path.splitext(name)[1].lower()
                if ext not in ALLOWED_EXT:
                    ext = '.bin'
                fname = '%s_%d%s' % (base, int(time.time() * 1000), ext)
                body = base64.b64decode(b64)
                # 存云存储（失败降级本地）
                ok = storage_put(UPLOAD_PREFIX + fname, body)
                if not ok:
                    os.makedirs(UPLOAD_DIR, exist_ok=True)
                    with open(os.path.join(UPLOAD_DIR, fname), 'wb') as f:
                        f.write(body)
                self._send_json({"ok": True, "url": '/uploads/' + fname, "name": fname})
            except Exception as e:
                self._send_json({"ok": False, "err": str(e)}, 400)
            return
        if self.path == '/api/export':
            try:
                length = int(self.headers.get('Content-Length', 0))
                raw = self.rfile.read(length)
                req = json.loads(raw.decode('utf-8'))
                rows = req.get('rows', [])
                if not isinstance(rows, list) or not rows:
                    raise ValueError('rows must be a non-empty array')
                body = build_import_file(rows)
                self.send_response(200)
                self.send_header('Content-Type', 'application/vnd.ms-excel; charset=utf-8')
                self.send_header('Content-Length', str(len(body)))
                self.send_header('Content-Disposition', 'attachment; filename="Vigor_export.xls"')
                self.end_headers()
                self.wfile.write(body)
                self.wfile.flush()
            except Exception as e:
                import traceback
                traceback.print_exc(file=sys.stderr)
                self._send_json({"ok": False, "err": str(e)}, 400)
            return
        if self.path == '/api/import':
            try:
                length = int(self.headers.get('Content-Length', 0))
                raw = self.rfile.read(length)
                req = json.loads(raw.decode('utf-8'))
                rows = req.get('rows', [])
                if not isinstance(rows, list) or not rows:
                    raise ValueError('rows must be a non-empty array')
                stats = do_import(rows)
                self._send_json({"ok": True, "stats": stats})
            except Exception as e:
                import traceback
                traceback.print_exc(file=sys.stderr)
                self._send_json({"ok": False, "err": str(e)}, 400)
            return
        if self.path == '/api/import_file':
            try:
                length = int(self.headers.get('Content-Length', 0))
                raw = self.rfile.read(length)
                req = json.loads(raw.decode('utf-8'))
                name = str(req.get('name', 'import.xlsx'))
                b64 = str(req.get('data', ''))
                if not b64:
                    raise ValueError('no file data')
                tmp = os.path.join(BASE, '_tmp_import.xlsx')
                with open(tmp, 'wb') as f:
                    f.write(base64.b64decode(b64))
                try:
                    wb = openpyxl.load_workbook(tmp)
                finally:
                    try:
                        if os.path.exists(tmp):
                            os.remove(tmp)
                    except Exception:
                        pass  # 沙箱环境可能禁止删除，临时文件同名覆盖即可
                ws = wb['产品导入'] if '产品导入' in wb.sheetnames else wb.active
                rows = []
                for r in range(2, ws.max_row + 1):
                    vals = [ws.cell(row=r, column=c).value for c in range(1, 28)]
                    if all(v is None or str(v).strip() == '' for v in vals):
                        continue
                    # 跳过模板黄色示例行及说明提示行，避免被导入为真实产品
                    cell1 = ws.cell(row=r, column=1)
                    try:
                        fl = cell1.fill
                        fg = str(fl.fgColor.rgb if fl and fl.fgColor else '')
                        if fl and fl.patternType and 'FEF3C7' in fg:
                            continue
                    except Exception:
                        pass
                    row_txt = ''.join(str(v) for v in vals if v is not None)
                    if '示例' in row_txt or '黄色' in row_txt or '删除后填写' in row_txt:
                        continue
                    def s(v):
                        return '' if v is None else str(v).strip()
                    params = []
                    for i in range(7):
                        # 参数列顺序：编码、名称(中)、名称(英)
                        code_txt = s(vals[6 + i * 3])
                        name_cn = s(vals[7 + i * 3])
                        name_en = s(vals[8 + i * 3])
                        if not (name_cn or name_en or code_txt):
                            continue
                        label = 'K%d' % (i + 1)
                        # 新格式：编码列 = 选项自定义编码，名称(中/英) = 选项描述
                        if ';' not in code_txt and '；' not in code_txt:
                            params.append({'label': label, 'code': code_txt, 'desc': name_cn, 'desc_en': name_en})
                            continue
                        # 旧格式兼容：编码列为 "编码 描述; 编码 描述"，名称列当组名
                        opts = []
                        for part in re.split(r'[;；]', code_txt):
                            part = part.strip()
                            if not part:
                                continue
                            seg = part.split(None, 1)
                            opts.append([seg[0], seg[1] if len(seg) > 1 else ''])
                        params.append({'label': label, 'name_cn': name_cn, 'name_en': name_en, 'opts': opts})
                    rows.append({
                        'api_cn': s(vals[0]), 'api_en': s(vals[1]),
                        'cat_cn': s(vals[2]), 'cat_en': s(vals[3]),
                        'prod_cn': s(vals[4]), 'prod_en': s(vals[5]),
                        'params': params,
                    })
                stats = do_import(rows)
                self._send_json({"ok": True, "file": name, "rows": len(rows), "stats": stats})
            except Exception as e:
                import traceback
                traceback.print_exc(file=sys.stderr)
                self._send_json({"ok": False, "err": str(e)}, 400)
            return
        self._send_json({"ok": False, "err": "not found"}, 404)

    def log_message(self, fmt, *args):
        pass


if __name__ == '__main__':
    os.makedirs(PUBLIC, exist_ok=True)
    # 诊断：打印环境变量键名（不打印值，避免泄露密钥）
    keys = sorted(k for k in os.environ if 'SECRET' in k.upper() or 'TENCENT' in k.upper() or k.startswith('TCB') or k == 'PORT' or 'TOKEN' in k.upper())
    print('ENV_KEYS: %s' % ','.join(keys), flush=True)
    print('PORT_ENV=%r -> bind=%d' % (os.environ.get('PORT'), PORT), flush=True)
    print('COS=%s bucket=%s region=%s' % ('enabled' if cos() is not None else 'disabled(本地磁盘模式)', COS_BUCKET, COS_REGION), flush=True)
    print('=' * 50, flush=True)
    socketserver.TCPServer.allow_reuse_address = True
    with socketserver.ThreadingTCPServer(('0.0.0.0', PORT), Handler) as httpd:
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print('\n服务已停止')
